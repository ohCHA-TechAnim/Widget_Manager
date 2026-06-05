"""넥슨 SharePoint Excel → Widget_Manager 일감 데이터 변환기.

TaskHub/utils/selenium_downloader.py 의 parse_excel_data 로직을 이식.
검증된 부분: 병합 셀 복제, 90% 공동 일감 파란색, 상태 접두사 감지.
새로 추가: 연속 날짜 → start/end 통합 (Widget_Manager 태스크 포맷).
"""
import logging
import re
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import urlparse

logger = logging.getLogger(__name__)


# ── 색상 헬퍼 (TaskHub math_utils 해당 로직 이식) ──────────────────────────────

def _safe_text(obj) -> str:
    if isinstance(obj, bytes):
        return obj.decode("utf-8", errors="ignore")
    return str(obj) if obj is not None else ""


def _normalize_argb(rgb_val) -> str:
    """openpyxl 셀 색상값을 8자리(AARRGGBB) hex로 정규화. 무효 시 ""."""
    if not rgb_val:
        return ""
    s = str(rgb_val)
    if len(s) == 8:
        return s
    if len(s) == 6:
        return "FF" + s
    return ""


def _argb_to_css(rgb_val) -> str:
    """8자리 ARGB → CSS '#RRGGBB'. 무효 시 ""."""
    n = _normalize_argb(rgb_val)
    return f"#{n[2:]}" if n else ""


def _ratio(part: int, whole: int) -> float:
    """0 나눗셈 방어 비율. whole이 0이면 0.0."""
    return part / whole if whole > 0 else 0.0


# ── 엑셀 시스템 텍스트 제거 ──────────────────────────────────────────────────────

# 엑셀 더미 URL 패턴 (go.microsoft.com/fwlink)
_EXCEL_DUMMY_URL_RE = re.compile(r'https?://go\.microsoft\.com/fwlink[^\s]*', re.IGNORECASE)

def _clean_excel_comment(text: str) -> str:
    """엑셀 스레드 댓글 안내문을 제거. 알려진 패턴만 제거, 나머지는 보존."""
    # [Threaded comment] 제거
    text = re.sub(r'\[Threaded comment\]\s*', '', text)
    # "Your version of Excel ... Learn more:" 패턴 제거 (한/영 혼용 대비 넉넉하게)
    text = re.sub(
        r'Your version of Excel[^.]*\..*?Learn more:\s*',
        '', text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    # go.microsoft.com/fwlink 더미 URL 제거
    text = _EXCEL_DUMMY_URL_RE.sub('', text)
    # 3줄 이상 연속 빈 줄 정리
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _infer_link_name(path: str) -> str:
    """URL/경로에서 표시용 이름 추론. 이름 지정이 없을 때 기본값으로 사용."""
    if not path:
        return ""
    if path.startswith(("http://", "https://")):
        try:
            parsed = urlparse(path)
            segments = [s for s in parsed.path.strip("/").split("/") if s]
            if segments:
                return segments[-1]
            return parsed.netloc
        except Exception:
            return path
    # 로컬 경로 → 파일/폴더명
    return Path(path).name or path


def _status_default_color(status: str) -> str:
    """상태별 기본 색상 (SharePoint 일감에 유효한 색이 없을 때 사용)."""
    if status == "doing":
        return "#C8803A"   # 중간 앰버 — 진행 중
    if status == "done":
        return "#7A7A8A"   # 차분한 회색 — 완료
    return "#5A8F6A"       # 중간 녹색 — 예정(todo/default)


# ── 상태·태그 파싱 ─────────────────────────────────────────────────────────────

def _detect_status(raw: str) -> str:
    """셀 접두사 기준으로 Widget_Manager status 추출."""
    if re.match(r'^진\)\s*', raw):
        return "doing"
    if re.match(r'^완\)\s*', raw):
        return "done"
    # 예), 보) → todo (예정 / 보류 모두 기본값)
    return "todo"


def _strip_prefix(text: str) -> str:
    """예)/완)/진)/보) 접두사 제거."""
    return re.sub(r'^(예\)|완\)|진\)|보\))\s*', '', text)


def _strip_release(text: str) -> tuple[str, str]:
    """릴리즈 태그(S1E1W2 형식)를 분리. (tag, remaining) 반환."""
    m = re.search(r'(S\d+(?:E\d+)?[\~_ -]*W\d+)', text, re.IGNORECASE)
    if m:
        tag = m.group(1).upper()
        rest = text.replace(m.group(0), "").strip()
        return tag, rest
    return "", text


# ── 메인 파서 ──────────────────────────────────────────────────────────────────

def parse_excel(
    excel_path: Path,
    target_name: str,
    sheet_name: str,
    target_year: int,
) -> list[dict]:
    """
    엑셀 파일을 파싱하여 Widget_Manager 일감 dict 목록을 반환한다.

    - source="sharepoint" 고정
    - 연속된 날짜(gap ≤ 3일)를 start/end로 통합
    - 90% 이상의 팀원이 동일 일감 → 색상 #00BFFF (TaskHub 검증 규칙 유지)
    - 셀 접두사(진)/완)) 로 status 감지
    """
    # str로 전달되는 경우를 방어적으로 처리
    excel_path = Path(excel_path)

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl이 설치되지 않음 — pip install openpyxl")
        return []

    if not excel_path.exists():
        logger.warning("SharePoint 엑셀 파일 없음: %s", excel_path)
        return []

    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    except Exception:
        logger.exception("엑셀 파일 열기 실패: %s", excel_path)
        return []

    # ── 시트 진단 로깅 ────────────────────────────────────────────────────
    logger.info("워크북 열기 성공: %s", excel_path.name)
    logger.info("시트 목록 (%d개): %s", len(wb.sheetnames), wb.sheetnames)

    # 시트명 불일치 시 첫 번째 시트로 폴백
    if sheet_name not in wb.sheetnames:
        fallback = wb.sheetnames[0] if wb.sheetnames else None
        if fallback is None:
            logger.error("시트 없음 — 파싱 중단")
            return []
        logger.warning(
            "설정 시트 '%s' 없음 → 첫 번째 시트 '%s'로 폴백. "
            "사용 가능한 시트: %s",
            sheet_name, fallback, wb.sheetnames,
        )
        sheet_name = fallback

    logger.info("사용 시트: '%s'", sheet_name)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    logger.info("시트 크기: %d행 × %d열", len(rows), len(rows[0]) if rows else 0)

    if len(rows) < 3:
        logger.error("엑셀 행 수 부족 (%d행) — 최소 3행 필요", len(rows))
        return []

    r1_cells = rows[0]   # 월 헤더 행
    r3_cells = rows[2]   # 날짜(일) 행

    # ── 대상자 행 탐색 ──────────────────────────────────────────────────────
    target_row_idx = -1
    name_col_idx = -1
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if _safe_text(cell.value).strip() == target_name:
                target_row_idx = i
                name_col_idx = j
                break
        if target_row_idx != -1:
            break

    if target_row_idx == -1:
        logger.error("대상자 '%s'를 시트에서 찾을 수 없음", target_name)
        return []

    target_cells = rows[target_row_idx]
    max_col = len(target_cells)

    # ── 팀원 행 목록 (90% / 80% 판정용) ────────────────────────────────────
    worker_rows: list[int] = []
    for r in range(3, len(rows)):
        if r == target_row_idx:
            worker_rows.append(r)
            continue
        if name_col_idx != -1 and name_col_idx < len(rows[r]):
            name_val = _safe_text(rows[r][name_col_idx].value).strip()
            if name_val and not name_val.isdigit() and len(name_val) < 15:
                worker_rows.append(r)

    # ── 병합 셀 룩업 테이블 (TaskHub 검증 패턴 그대로) ──────────────────────
    merged_lookup: dict = {}
    try:
        for crange in ws.merged_cells.ranges:
            try:
                clo, rlo, chi, rhi = crange.bounds
                top_cell = ws.cell(row=rlo, column=clo)
                top_val = top_cell.value
                top_comment = top_cell.comment
                top_hyperlink = top_cell.hyperlink
                top_fill = top_cell.fill
                for r in range(rlo, rhi + 1):
                    for c in range(clo, chi + 1):
                        merged_lookup[(r, c)] = (top_val, top_comment, top_hyperlink, top_fill)
            except Exception:
                logger.warning("병합 셀 처리 오류 (건너뜀): %s", crange, exc_info=True)
    except Exception:
        logger.warning("병합 셀 전체 처리 오류 — 병합 셀 없이 계속", exc_info=True)
    logger.info("병합 셀 룩업 구축 완료: %d개 항목", len(merged_lookup))

    # ── 날짜별 일감 수집 ────────────────────────────────────────────────────
    day_entries: list[dict] = []
    active_task: str | None = None
    active_color = ""
    active_status = "todo"
    active_release = ""
    active_memo = ""
    active_jiras: list[str] = []
    month_matched: int | None = None

    _skipped_cols = 0
    for col_idx in range(len(r3_cells)):
        try:
            c1 = r1_cells[col_idx] if col_idx < len(r1_cells) else None
            c3 = r3_cells[col_idx] if col_idx < len(r3_cells) else None
            c1_val = _safe_text(c1.value) if c1 else ""
            c3_val = _safe_text(c3.value) if c3 else ""

            # 월 헤더 감지
            if c1_val:
                m = re.search(r'(\d+)월', c1_val)
                if m:
                    month_matched = int(m.group(1))

            if not month_matched or not c3_val:
                continue

            m_day = re.search(r'^(\d+)', c3_val)
            if not m_day:
                continue

            d = int(m_day.group(1))
            try:
                date_obj = date(target_year, month_matched, d)
            except ValueError:
                logger.debug("날짜 파싱 실패 — %d월 %d일 (건너뜀)", month_matched, d)
                continue

            # 대상자 셀 읽기 (병합 셀 우선)
            tc_row = target_row_idx + 1
            tc_col = col_idx + 1
            if (tc_row, tc_col) in merged_lookup:
                m_val, m_comment, m_hyperlink, m_fill = merged_lookup[(tc_row, tc_col)]
                tc_val = _safe_text(m_val)
                tc_comment = m_comment
                tc_hyperlink = m_hyperlink
                tc_fill = m_fill
            else:
                tc = target_cells[col_idx] if col_idx < max_col else None
                tc_val = _safe_text(tc.value) if tc else ""
                tc_comment = tc.comment if tc else None
                tc_hyperlink = tc.hyperlink if tc else None
                tc_fill = tc.fill if tc else None

            # 셀 내용 파싱
            if tc_val:
                t_raw = tc_val.strip()
                if t_raw and t_raw.lower() not in ("none", "-"):
                    active_status = _detect_status(t_raw)
                    t_clean = _strip_prefix(t_raw)
                    active_release, t_clean = _strip_release(t_clean)
                    t_clean = t_clean.strip()
                    active_task = t_clean if t_clean else None

                    # 색상 — GradientFill 등 예상치 못한 fill 타입에 방어적으로 접근
                    active_color = ""
                    try:
                        if tc_fill and hasattr(tc_fill, "start_color") and tc_fill.start_color:
                            rgb_val = getattr(tc_fill.start_color, "rgb", None)
                            if rgb_val:
                                active_color = _argb_to_css(rgb_val)
                    except Exception:
                        pass  # 색상 읽기 실패는 무시

                    # 코멘트 → 메모 + Jira
                    active_memo = ""
                    active_jiras = []
                    if tc_comment and getattr(tc_comment, "text", None):
                        cmt = _clean_excel_comment(_safe_text(tc_comment.text))
                        for url in re.findall(r"https?://[^\s]+", cmt):
                            # go.microsoft.com/fwlink 더미 URL 제외
                            if "go.microsoft.com/fwlink" not in url and url not in active_jiras:
                                active_jiras.append(url)
                        cmt_clean = re.sub(r"https?://[^\s]+", "", cmt).strip()
                        if cmt_clean:
                            active_memo = cmt_clean

                    # 셀 실제 하이퍼링크 우선 추출 (더미 링크 제외)
                    if tc_hyperlink and getattr(tc_hyperlink, "target", None):
                        tgt = _safe_text(tc_hyperlink.target).strip()
                        if (tgt
                                and "go.microsoft.com/fwlink" not in tgt
                                and tgt not in active_jiras):
                            active_jiras.append(tgt)
                else:
                    active_task = None
            # tc_val이 빈 경우: active_task 유지 (뒤에 tc_val 조건으로 기록 여부 결정)

            # 실제 내용이 있는 셀만 기록 (빈 셀 = 주말/공백은 건너뜀)
            if not active_task or not tc_val:
                continue

        except Exception:
            _skipped_cols += 1
            logger.warning("열 %d 처리 중 예외 — 건너뜀", col_idx, exc_info=True)
            continue

        # 90% 공동 일감 판정 (TaskHub 검증 로직 유지)
        match_count = 0
        for r in worker_rows:
            if (r + 1, col_idx + 1) in merged_lookup:
                w_val = _safe_text(merged_lookup[(r + 1, col_idx + 1)][0]).strip()
            else:
                w_val = _safe_text(rows[r][col_idx].value).strip() if col_idx < len(rows[r]) else ""
            w_clean = _strip_prefix(w_val)
            _, w_clean = _strip_release(w_clean)
            w_clean = w_clean.strip()
            if w_clean == active_task:
                match_count += 1

        # 전체 팀원의 90% 이상 동일 일감 → 공유 공동 일감 #00BFFF
        is_shared = _ratio(match_count, len(worker_rows)) >= 0.9
        if is_shared:
            final_color = "#00BFFF"
        elif active_color in ("", "#000000", "#FFFFFF"):
            # 유효한 셀 색상 없음 → 상태별 기본색 적용
            final_color = _status_default_color(active_status)
        else:
            final_color = active_color

        day_entries.append({
            "date": date_obj,
            "task": active_task,
            "color": final_color,
            "status": active_status,
            "release": active_release,
            "memo": active_memo,
            "jiras": list(active_jiras),
        })

    if _skipped_cols:
        logger.warning("열 처리 중 오류로 건너뜀: %d개 열", _skipped_cols)

    if not day_entries:
        logger.info("파싱 결과 없음 (대상자: %s, 시트: %s)", target_name, sheet_name)
        return []

    # ── 연속 날짜 → start/end 통합 ──────────────────────────────────────────
    # 동일 task, gap ≤ 3일(주말 포함) → 하나의 일감으로 통합
    day_entries.sort(key=lambda x: x["date"])

    groups: list[list[dict]] = []
    current_group = [day_entries[0]]
    for entry in day_entries[1:]:
        prev = current_group[-1]
        gap = (entry["date"] - prev["date"]).days
        if entry["task"] == prev["task"] and gap <= 3:
            current_group.append(entry)
        else:
            groups.append(current_group)
            current_group = [entry]
    groups.append(current_group)

    result: list[dict] = []
    for group in groups:
        first = group[0]
        last = group[-1]

        memo = first["memo"]
        if first["release"]:
            memo = (f"[릴리즈: {first['release']}]\n" + memo).strip()

        jiras = [{"name": _infer_link_name(j), "path": j} for j in first["jiras"]]

        result.append({
            "title": first["task"],
            "start": first["date"].isoformat(),
            "end": last["date"].isoformat(),
            "status": first["status"],
            "priority": "mid",
            "color": first["color"],
            "memo": memo,
            "jiras": jiras,
            "folders": [],
            "attachments": [],
            "source": "sharepoint",
        })
        logger.debug(
            "파싱: %s  %s ~ %s  (%s)",
            first["task"], first["date"], last["date"], first["status"],
        )

    logger.info("SharePoint 파싱 완료: %d개 일감 (대상자: %s)", len(result), target_name)
    return result
