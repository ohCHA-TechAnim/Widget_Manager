"""SharePoint 파서 단위 테스트 — 네트워크·실제 엑셀 불필요.

헬퍼 함수 검증 + 샘플 엑셀(openpyxl로 인메모리 생성)로 end-to-end 파싱 검증.
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from plugins.nexon_sharepoint.parser import (
    _detect_status,
    _strip_prefix,
    _strip_release,
    _normalize_argb,
    _argb_to_css,
    _ratio,
    parse_excel,
)


# ── 헬퍼 함수 ────────────────────────────────────────────────────────────────

def test_detect_status_doing():
    assert _detect_status("진) 캐릭터A 작업") == "doing"


def test_detect_status_done():
    assert _detect_status("완) 배경 작업") == "done"


def test_detect_status_todo():
    assert _detect_status("예) 기획") == "todo"
    assert _detect_status("보) 대기") == "todo"
    assert _detect_status("일반 작업") == "todo"


def test_strip_prefix():
    assert _strip_prefix("진) 캐릭터A") == "캐릭터A"
    assert _strip_prefix("완) 배경") == "배경"
    assert _strip_prefix("예) 기획") == "기획"
    assert _strip_prefix("일반") == "일반"


def test_strip_release_found():
    tag, rest = _strip_release("캐릭터A S1E1W2")
    assert tag == "S1E1W2"
    assert rest.strip() == "캐릭터A"


def test_strip_release_not_found():
    tag, rest = _strip_release("일반 작업")
    assert tag == ""
    assert rest == "일반 작업"


def test_normalize_argb_8digit():
    assert _normalize_argb("FF00BFFF") == "FF00BFFF"


def test_normalize_argb_6digit():
    assert _normalize_argb("00BFFF") == "FF00BFFF"


def test_normalize_argb_none():
    assert _normalize_argb(None) == ""
    assert _normalize_argb("") == ""


def test_argb_to_css():
    assert _argb_to_css("FF00BFFF") == "#00BFFF"
    assert _argb_to_css("FFFF9494") == "#FF9494"
    assert _argb_to_css(None) == ""


def test_ratio():
    assert _ratio(9, 10) == pytest.approx(0.9)
    assert _ratio(0, 0) == 0.0
    assert _ratio(8, 10) == pytest.approx(0.8)
    assert _ratio(1, 1) == pytest.approx(1.0)


# ── parse_excel — 파일 없음 케이스 ───────────────────────────────────────────

def test_parse_excel_missing_file():
    """존재하지 않는 엑셀은 빈 리스트를 반환해야 한다."""
    result = parse_excel(
        excel_path=Path("/nonexistent/file.xlsx"),
        target_name="테스트",
        sheet_name="애니메이션",
        target_year=2026,
    )
    assert result == []


# ── 샘플 엑셀 파싱 end-to-end ────────────────────────────────────────────────

def _build_sample_excel(tmp_path: Path) -> Path:
    """
    최소 구조의 샘플 엑셀 생성.

    Row 1: [이름열] [1월] [빈칸...x6]
    Row 2: (비어있음)
    Row 3: [빈칸]  [1]  [2]  [3]  [4]  [5]  [6]  [7]
    Row 4: [테스트] [캐릭터A] [캐릭터A] [진) 배경B] [진) 배경B] [None] [None] [None]
    """
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl 미설치 — 샘플 엑셀 테스트 건너뜀")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "애니메이션"

    # Row 1: 월 헤더
    ws.cell(1, 1, "이름")
    ws.cell(1, 2, "1월")

    # Row 3: 날짜(일) 헤더
    ws.cell(3, 1, "")
    for col, day in enumerate(range(1, 8), start=2):
        ws.cell(3, col, str(day))

    # Row 4: 대상자 일감
    ws.cell(4, 1, "테스트")
    ws.cell(4, 2, "캐릭터A")
    ws.cell(4, 3, "캐릭터A")
    ws.cell(4, 4, "진) 배경B")
    ws.cell(4, 5, "진) 배경B")
    # 5, 6, 7열은 비어 있음 (주말)

    out = tmp_path / "sample.xlsx"
    wb.save(str(out))
    return out


def test_parse_sample_excel(tmp_path):
    """샘플 엑셀 파싱 결과 구조 검증."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl 미설치")

    excel_path = _build_sample_excel(tmp_path)
    tasks = parse_excel(
        excel_path=excel_path,
        target_name="테스트",
        sheet_name="애니메이션",
        target_year=2026,
    )

    assert isinstance(tasks, list)
    assert len(tasks) >= 1, "최소 1개 이상의 일감이 파싱되어야 한다"

    for t in tasks:
        # 필수 필드 존재
        assert "title" in t
        assert "start" in t
        assert "end" in t
        assert "status" in t
        assert "source" in t
        # source 고정값
        assert t["source"] == "sharepoint"
        # status 유효값
        assert t["status"] in ("todo", "doing", "done")
        # start ≤ end
        assert t["start"] <= t["end"]
        # priority 기본값
        assert t.get("priority") == "mid"

    titles = {t["title"]: t for t in tasks}

    if "캐릭터A" in titles:
        assert titles["캐릭터A"]["status"] == "todo", "캐릭터A는 접두사 없음 → todo"
        assert titles["캐릭터A"]["start"] == "2026-01-01"
        assert titles["캐릭터A"]["end"] == "2026-01-02"

    if "배경B" in titles:
        assert titles["배경B"]["status"] == "doing", "진) 접두사 → doing"
        assert titles["배경B"]["start"] == "2026-01-03"
        assert titles["배경B"]["end"] == "2026-01-04"


def test_parse_target_not_found(tmp_path):
    """대상자가 시트에 없으면 빈 리스트를 반환해야 한다."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl 미설치")

    excel_path = _build_sample_excel(tmp_path)
    result = parse_excel(
        excel_path=excel_path,
        target_name="존재하지않는사람",
        sheet_name="애니메이션",
        target_year=2026,
    )
    assert result == []
